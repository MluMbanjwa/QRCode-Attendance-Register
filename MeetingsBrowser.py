import sys
import sqlite3
from PyQt6.QtWidgets import (
    QApplication, QTableWidget, QTableWidgetItem,
    QVBoxLayout, QHBoxLayout, QWidget, QComboBox,QPushButton,QFileDialog
)
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

bold_font = Font(bold=True)

class TableViewer(QWidget):
    def __init__(self):
        super().__init__()
        # Open database
        self.filename=''
        self.workbook = Workbook()
        self.conn = sqlite3.connect("DataBase.db")
        self.cursor = self.conn.cursor()
        
        # Window setup
        self.setWindowTitle("Past Meetings Attendance")
        self.resize(600, 500)
        
        # Comboboxes_buttns
        self.selectDate = QComboBox(self)
        self.selectTime = QComboBox(self)
        self.selectName=QComboBox(self)
        self.btnPrint=QPushButton('Export',self)
        self.addNames()
        self.addDates(self.selectName.currentText())
        self.addTime()
        
        # Connect signals
        #self.selectDate.currentTextChanged.connect(self.addTime)
        #self.selectName.currentTextChanged.connect(self.addDates)
        
        self.selectTime.currentTextChanged.connect(self.load_data)
        self.selectDate.currentTextChanged.connect(self.load_data)
        self.selectName.currentTextChanged.connect(self.load_data)
        
        self.btnPrint.clicked.connect(self.export)
        
        # Layouts
        comboboxLayout = QHBoxLayout()
        layout = QVBoxLayout(self)
        self.table = QTableWidget()

        comboboxLayout.addWidget(self.selectName)
        comboboxLayout.addWidget(self.selectDate)
        comboboxLayout.addWidget(self.selectTime)
        comboboxLayout.addWidget(self.btnPrint)
        layout.addLayout(comboboxLayout)
        layout.addWidget(self.table)
        self.setLayout(layout)
        self.load_data()
        
        self.show()

    def load_data(self):
        query=""
        if self.selectName.currentText()=="Name" and self.selectDate.currentText()=="Date" and self.selectTime.currentText()=="Time":
            query = """
        SELECT Meeting_Date,Meeting_Time,Meeting_Name, Teacher_Name, Teacher_Arrival_Time 
        FROM Data 
        """
            self.cursor.execute(query)
        elif self.selectName.currentText()!="Name" and self.selectDate.currentText()=="Date" and self.selectTime.currentText()=="Time":
            query = """
        SELECT Meeting_Date,Meeting_Time,Meeting_Name,Teacher_Name, Teacher_Arrival_Time 
        FROM Data 
        WHERE Meeting_Name = ?
        """
            self.cursor.execute(query, (self.selectName.currentText(),))
        elif self.selectName.currentText()=="Name" and self.selectDate.currentText()!="Date" and self.selectTime.currentText()=="Time":
            query = """
        SELECT Meeting_Date,Meeting_Time,Meeting_Name,Teacher_Name, Teacher_Arrival_Time 
        FROM Data 
        WHERE Meeting_Date = ?
        """
            self.cursor.execute(query, (self.selectDate.currentText(),))
        elif self.selectName.currentText()=="Name" and self.selectDate.currentText()=="Date" and self.selectTime.currentText()!="Time":
            query = """
        SELECT Meeting_Date,Meeting_Time,Meeting_Name,Teacher_Name, Teacher_Arrival_Time 
        FROM Data 
        WHERE Meeting_Time = ?
        """
            self.cursor.execute(query, (self.selectTime.currentText(),))
        elif self.selectName.currentText()=="Name" and self.selectDate.currentText()!="Date" and self.selectTime.currentText()!="Time":
            query = """
        SELECT Meeting_Date,Meeting_Time,Meeting_Name,Teacher_Name, Teacher_Arrival_Time 
        FROM Data 
        WHERE Meeting_Date = ? AND Meeting_Time = ?
        """
            self.cursor.execute(query, (self.selectDate.currentText(),self.selectTime.currentText()))
        elif self.selectName.currentText()!="Name" and self.selectDate.currentText()!="Date" and self.selectTime.currentText()=="Time":
            query = """
        SELECT Meeting_Date,Meeting_Time,Meeting_Name,Teacher_Name, Teacher_Arrival_Time 
        FROM Data 
        WHERE Meeting_Name = ? AND Meeting_Date = ?
        """
            self.cursor.execute(query, (self.selectName.currentText(), self.selectDate.currentText()))
        elif self.selectName.currentText()!="Name" and self.selectDate.currentText()=="Date" and self.selectTime.currentText()!="Time":
            query = """
        SELECT Meeting_Date,Meeting_Time,Meeting_Name,Teacher_Name, Teacher_Arrival_Time 
        FROM Data 
        WHERE Meeting_Name = ? AND Meeting_Time = ?
        """
            self.cursor.execute(query, (self.selectName.currentText(), self.selectTime.currentText()))
        rows = self.cursor.fetchall()

        self.table.setRowCount(len(rows))
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["Date","Start Time","Meeting Name","Teacher Name", "Arrival Time"])

        for row_idx, row_data in enumerate(rows):
            for col_idx, item in enumerate(row_data):
                self.table.setItem(row_idx, col_idx, QTableWidgetItem(str(item)))

    def addDates(self,names:str):
        self.selectDate.clear()
        if names=="Name":
            query = "SELECT DISTINCT Meeting_Date FROM Data ORDER BY Meeting_Date ASC"
            self.cursor.execute(query)
        else:
            query = "SELECT DISTINCT Meeting_Date FROM Data WHERE Meeting_Name=? ORDER BY Meeting_Date ASC"
            self.cursor.execute(query,(names,))
        dates = self.cursor.fetchall()
        self.selectDate.addItem('Date')
        for date_tuple in dates:
            self.selectDate.addItem(date_tuple[0])
        
    def addNames(self):
        self.selectName.clear()
        self.cursor.execute("SELECT DISTINCT Meeting_Name FROM Data ORDER BY Meeting_Date ASC")
        names = self.cursor.fetchall()
        self.selectName.addItem("Name")
        for name_tuple in names:
            self.selectName.addItem(name_tuple[0])
            
    def addTime(self):
        self.selectTime.clear()
        self.selectTime.addItem('Time')
        date=self.selectDate.currentText()
        if date!="Date":
            query = "SELECT DISTINCT Meeting_Time FROM Data WHERE Meeting_Date = ? ORDER BY Meeting_Time ASC"
            self.cursor.execute(query, (date,))
        else:
            query = "SELECT DISTINCT Meeting_Time FROM Data ORDER BY Meeting_Time ASC"
            self.cursor.execute(query)
        times = self.cursor.fetchall()
        for time_tuple in times:
            self.selectTime.addItem(time_tuple[0])    
    def export(self):
    # Ask user where to save the file
        self.workbook = Workbook()
        if self.filename=='':
            self.filename, _ = QFileDialog.getSaveFileName(self, "Save File", "", "Excel Files (*.xlsx)")
        sheet = self.workbook.active
        sheet.title = f"E-Register"

        #excell tittles
        sheet.cell(1, 2,f"Electronic Attendance Register 1.2:").font=bold_font
        sheet.cell(2, 2,f"Letsatsing Combined School").font=bold_font
        sheet.cell(3, 2,f"Attendance Register").font=bold_font
        sheet.column_dimensions[get_column_letter(1)].width = 5 #index
        for x in range(2,6):
            sheet.column_dimensions[get_column_letter(x)].width = 30
        sheet.column_dimensions[get_column_letter(6)].width = 20
        # Write headers
        for col in range(self.table.columnCount()):
            header = self.table.horizontalHeaderItem(col).text()
            sheet.cell(row=1+3, column=col + 2, value=header).font=bold_font

        # Write data
        for row in range(self.table.rowCount()):
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                sheet.cell(row=row + 2+3, column=col + 2, value=item.text() if item else "")
                sheet.cell(row=row + 2+3, column=1, value=str(row+1) if item else "")

        # Save to file
        self.workbook.save(self.filename)
